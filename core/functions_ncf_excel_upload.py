from openpyxl import load_workbook
from datetime import datetime
import xlrd, datetime as DT, decimal, numbers
from django.core.exceptions import ObjectDoesNotExist

from .functions_shared import app_get_agreement_id

from .models import ncf_props_and_payouts as PropsFile, \
    ncf_settled_agreements as SettledFile, \
    ncf_advanced_payments, \
    ncf_arrears_collected, \
    ncf_agreement_titles, \
    ncf_global_terminations


def app_process_props_files():
    PropsFile.objects.all().delete()

    filepath = "//ncfsrv01-mk/Company/Accounts - Marc/Account B/09 - Alfandari Private Equities Ltd/Own Book/AOB Props & Payouts.xlsx"
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    index = 460
    val_agreement_id = ws.cell(row=index, column=6).value

    # iterate over all rows
    while val_agreement_id:

        val_regulated_flag = ws.cell(row=index, column=4).value
        val_payout_date = ws.cell(row=index, column=5).value
        val_customer_name = ws.cell(row=index, column=7).value
        val_sales_person = ws.cell(row=index, column=8).value
        val_rep_person = ws.cell(row=index, column=9).value
        val_term_text = ws.cell(row=index, column=13).value
        val_term_mm = ws.cell(row=index, column=14).value
        val_net_invoice_amount = ws.cell(row=index, column=15).value
        val_net_gross_amount = ws.cell(row=index, column=16).value
        val_agreement_type = ws.cell(row=index, column=18).value
        val_first_rental_date = ws.cell(row=index, column=32).value
        val_final_rental_date = ws.cell(row=index, column=33).value
        val_gross_rental = ws.cell(row=index, column=25).value

        if type(val_payout_date) is not datetime:
            val_payout_date = '2000-01-01'

        if type(val_first_rental_date) is not datetime:
            val_first_rental_date = '2000-01-01'

        if type(val_final_rental_date) is not datetime:
            val_final_rental_date = '2000-01-01'

        # write new row
        new_props = PropsFile.objects.create(agreement_id=val_agreement_id,
                                             regulated_flag=val_regulated_flag,
                                             payout_date=val_payout_date,
                                             customer_name=val_customer_name,
                                             sales_person=val_sales_person,
                                             rep_person=val_rep_person,
                                             term_text=val_term_text,
                                             term_mm=val_term_mm,
                                             gross_rental=val_gross_rental,
                                             net_invoice_amount=val_net_invoice_amount,
                                             net_gross_amount=val_net_gross_amount,
                                             agreement_type=val_agreement_type,
                                             first_rental_date=val_first_rental_date,
                                             final_rental_date=val_final_rental_date)

        index += 1
        val_agreement_id = ws.cell(row=index, column=6).value


def app_process_settled_agreements():
    SettledFile.objects.all().delete()

    filepath = "//ncfsrv01-mk/Company/AOB End Of Day Report/Arrears Report.2019.email.xls"
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(1)

    # iterate over all rows
    for index in range(1, ws.nrows):

        val_agreement_id = ws.cell(index, 0).value
        val_agreement_name = ws.cell(index, 1).value

        try:
            float(ws.cell(index, 2).value)
            val_settlement_value = ws.cell(index, 2).value
        except ValueError:
            val_settlement_value = 0

        try:
            val_method = str(int(ws.cell(index, 3).value))
        except:
            val_method = ws.cell(index, 3).value

        try:
            wip_settlement_date = int(ws.cell(index, 4).value)
            year, month, day, hour, minute, second = xlrd.xldate_as_tuple(wip_settlement_date, wb.datemode)
            val_settlement_date = DT.datetime(year, month, day, hour, minute, second)
        except:
            val_settlement_date = '2000-01-01'

        val_notes = ws.cell(index, 5).value
        val_received_from = ws.cell(index, 6).value
        val_agreement_type = ws.cell(index, 7).value
        val_vat_status = ws.cell(index, 8).value
        val_removed_from_sentinel = ws.cell(index, 9).value

        if type(val_removed_from_sentinel) is not bool:
            val_removed_from_sentinel = False

        # write new row
        new_settled = SettledFile.objects.create(agreement_id=val_agreement_id,
                                                 agreement_name=val_agreement_name,
                                                 settlement_value=val_settlement_value,
                                                 method=val_method,
                                                 settlement_date=val_settlement_date,
                                                 notes=val_notes,
                                                 received_from=val_received_from,
                                                 agreement_type=val_agreement_type,
                                                 vat_status=val_vat_status,
                                                 removed_from_sentinel=val_removed_from_sentinel)


def app_process_advance_payments():
    ncf_advanced_payments.objects.all().delete()

    filepath = "//ncfsrv01-mk/Company/AOB End Of Day Report/Arrears Report.2019.email.xls"
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    save_advance = '2018-01-01'
    wip_agreement_id = ' '

    # iterate over all rows
    for index in range(7, ws.nrows):

        # get date
        try:
            wip_advance_date = int(ws.cell(index, 10).value)
            year, month, day, hour, minute, second = xlrd.xldate_as_tuple(wip_advance_date, wb.datemode)
            save_advance_date = DT.datetime(year, month, day, hour, minute, second)
        except:
            search_reference = ws.cell(index, 10).value
            wip_agreement_id = app_get_agreement_id(search_reference)


        # process advance details
        try:
            props_and_payouts_row = PropsFile.objects.get(agreement_id=wip_agreement_id)

            val_agreement_id = ws.cell(index, 10).value
            val_agreement_name = ws.cell(index, 11).value
            val_advance_value = ws.cell(index, 12).value
            val_advance_date = save_advance_date
            val_notes = ws.cell(index, 14).value

            try:
                val_method = str(int(ws.cell(index, 13).value))
            except:
                val_method = ws.cell(index, 13).value

            # write new row
            new_advanced = ncf_advanced_payments.objects.create(agreement_id=val_agreement_id,
                                                                agreement_name=val_agreement_name,
                                                                advance_value=val_advance_value,
                                                                method=val_method,
                                                                advance_date=val_advance_date,
                                                                notes=val_notes)

        except ObjectDoesNotExist:
            pass
        else:
            pass


def app_process_arrears_collected():
    ncf_arrears_collected.objects.all().delete()

    filepath = "//ncfsrv01-mk/Company/AOB End Of Day Report/Arrears Report.2019.email.xls"
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    save_advance = '2018-01-01'
    wip_agreement_id = ' '

    # iterate over all rows
    for index in range(7, ws.nrows):


        # get date
        try:
            wip_advance_date = int(ws.cell(index, 1).value)
            year, month, day, hour, minute, second = xlrd.xldate_as_tuple(wip_advance_date, wb.datemode)
            save_collected_date = DT.datetime(year, month, day, hour, minute, second)
        except:
            try:
                wip_advance_date = datetime.strptime(ws.cell(index, 1).value, "%d/%m/%Y")
                save_collected_date = wip_advance_date
            except:
                try:
                    wip_advance_date = datetime.strptime(ws.cell(index, 1).value, "%d/%m/%y")
                    save_collected_date = wip_advance_date
                except:
                    search_reference = ws.cell(index, 1).value
                    wip_agreement_id = app_get_agreement_id(search_reference)


        # process advance details
        try:
            props_and_payouts_row = PropsFile.objects.get(agreement_id=wip_agreement_id)

            val_collected_date = save_collected_date
            val_agreement_id = ws.cell(index, 1).value
            val_agreement_name = ws.cell(index, 2).value

            if isinstance(ws.cell(index, 3).value, numbers.Number):
                val_collected_value = ws.cell(index, 3).value
            else:
                val_collected_value = 0
            val_method = ws.cell(index, 5).value
            if isinstance(ws.cell(index, 6).value, decimal.Decimal):
                val_fees = ws.cell(index, 6).value
            else:
                val_fees = 0
            val_agent_name = ws.cell(index, 7).value
            val_notes = ws.cell(index, 8).value

            try:
                val_method = str(int(ws.cell(index, 5).value))
            except:
                val_method = ws.cell(index, 5).value

            try:
                val_num_check = int(val_fees)
            except:
                val_fees = 0

            # write new row
            new_collected = ncf_arrears_collected.objects.create(ac_collected_date=val_collected_date,
                                                                 ac_agreement_id=val_agreement_id,
                                                                 ac_agreement_name=val_agreement_name,
                                                                 ac_arrears_collected=val_collected_value,
                                                                 ac_method=val_method,
                                                                 ac_fees_collected=val_fees,
                                                                 ac_agent_name=val_agent_name,
                                                                 ac_notes=val_notes
                                                                 )
        except ObjectDoesNotExist:
            pass
        else:
            pass


def app_process_title_payments():
    ncf_agreement_titles.objects.all().delete()

    filepath = "C:/Users/Steve.Ismay/Desktop/Live Title Report.xls"
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    save_title_date = '2018-01-01'
    wip_agreement_id = ' '

    # iterate over all rows
    for index in range(6, ws.nrows):

        row_type = ''

        wip_agreement_id = ws.cell(index, 3).value
        val_agreement_id = wip_agreement_id.strip()
        if val_agreement_id != '':
            row_type = 'agreement'
        else:
            try:
                wip_title_date = int(ws.cell(index, 4).value)
                year, month, day, hour, minute, second = xlrd.xldate_as_tuple(wip_title_date, wb.datemode)
                save_title_date = DT.datetime(year, month, day, hour, minute, second)
                row_type = 'date'
            except:
                pass

        if row_type == 'agreement':
            val_title_date = save_title_date
            val_invoice_number = ws.cell(index, 4).value
            val_customer_name = ws.cell(index, 5).value

            try:
                float(ws.cell(index, 7).value)
                val_amount_paid = ws.cell(index, 7).value
            except:
                val_amount_paid = 0

            try:
                float(ws.cell(index, 8).value)
                val_paying_value = ws.cell(index, 8).value
            except:
                val_paying_value = 0

            try:
                float(ws.cell(index, 9).value)
                val_proforma = ws.cell(index, 9).value
            except:
                val_proforma = 0

            val_method = ws.cell(index, 13).value
            val_notes = ws.cell(index, 14).value
            val_by = ws.cell(index, 15).value
            # write new row
            new_title = ncf_agreement_titles.objects.create(agreement_id=val_agreement_id,
                                                            title_date=val_title_date,
                                                            invoice_number=val_invoice_number,
                                                            customer_name=val_customer_name,
                                                            amount_paid=val_amount_paid,
                                                            paying_value=val_paying_value,
                                                            proforma=val_proforma,
                                                            method=val_method,
                                                            notes=val_notes,
                                                            by=val_by)


def app_global_terminations():

    ncf_global_terminations.objects.all().delete()

    # Worksheet Name = Live
    filepath = "//ncfsrv01-mk/Company/Recoveries/AOB Global Termination List Master.xlsx"
    wb = load_workbook(filepath, data_only=True)
    ws = wb.get_sheet_by_name(name='Live')
    ws1 = wb.get_sheet_by_name(name='Terminated (With Loss)')
    ws2 = wb.get_sheet_by_name(name='Terminated (No Loss)')
    ws3 = wb.get_sheet_by_name(name='Closed')

    # get max row count
    max_row = ws.max_row
    max_row1 = ws1.max_row
    max_row2 = ws2.max_row
    max_row3 = ws3.max_row
    val_reason = 'TERMINATED - Worksheet = Live'
    val_reason1 = 'TERMINATED - Worksheet = Terminated (With Loss)'
    val_reason2 = 'TERMINATED - Worksheet = Terminated (No Loss)'
    val_reason3 = 'TERMINATED - Worksheet = Closed'

    for index in range(3, max_row + 1):

        wip_agreement_id = ws.cell(index, 1).value

        if wip_agreement_id and wip_agreement_id != '':

            val_agreement_id = wip_agreement_id.strip()
            val_agreement_name = ws.cell(row=index, column=2).value
            val_agreement_rep = ws.cell(row=index, column=3).value

            if ws.cell(row=index, column=5).value == 'Yes':
                val_written_off = True
            else:
                val_written_off = False

            val_date_terminated = ws.cell(row=index, column=6).value
            if type(val_date_terminated) is not datetime:
                val_date_terminated = '2000-01-01'

            # write new row
            new_global_termination = ncf_global_terminations.objects.create(agreement_id=val_agreement_id,
                                                                       agreement_name=val_agreement_name,
                                                                       agreement_rep=val_agreement_rep,
                                                                       written_off=val_written_off,
                                                                       date_terminated=val_date_terminated,
                                                                       reason=val_reason
                                                                       )

        # elif ws.cell(row=index, column=4).value:
        #
        #     if ws.cell(row=index, column=4).value.strip() != '':
        #         val_reason = ws.cell(row=index, column=4).value
        #     else:
        #         val_reason = 'TERMINATED - Worksheet = Live'

    for index in range(3, max_row1 + 1):

        wip_agreement_id = ws1.cell(index, 1).value

        if wip_agreement_id and wip_agreement_id != '':

            val_agreement_id = wip_agreement_id.strip()
            val_agreement_name = ws1.cell(row=index, column=2).value
            val_agreement_rep = ws1.cell(row=index, column=3).value

            if ws1.cell(row=index, column=5).value == 'Yes':
                val_written_off = True
            else:
                val_written_off = False

            val_date_terminated = ws1.cell(row=index, column=6).value
            if type(val_date_terminated) is not datetime:
                val_date_terminated = '2000-01-01'

            # write new row
            new_global_termination = ncf_global_terminations.objects.create(agreement_id=val_agreement_id,
                                                                       agreement_name=val_agreement_name,
                                                                       agreement_rep=val_agreement_rep,
                                                                       written_off=val_written_off,
                                                                       date_terminated=val_date_terminated,
                                                                       reason=val_reason1
                                                                       )

        # elif ws.cell(row=index, column=4).value:
        #
        #     if ws.cell(row=index, column=4).value.strip() != '':
        #         val_reason = ws.cell(row=index, column=4).value
        #     else:
        #         val_reason = 'TERMINATED - Worksheet = Live'

    for index in range(3, max_row2 + 1):

        wip_agreement_id = ws2.cell(index, 1).value

        if wip_agreement_id and wip_agreement_id != '':

            val_agreement_id = wip_agreement_id.strip()
            val_agreement_name = ws2.cell(row=index, column=2).value
            val_agreement_rep = ws2.cell(row=index, column=3).value

            if ws2.cell(row=index, column=5).value == 'Yes':
                val_written_off = True
            else:
                val_written_off = False

            val_date_terminated = ws2.cell(row=index, column=6).value
            if type(val_date_terminated) is not datetime:
                val_date_terminated = '2000-01-01'

            # write new row
            new_global_termination = ncf_global_terminations.objects.create(agreement_id=val_agreement_id,
                                                                       agreement_name=val_agreement_name,
                                                                       agreement_rep=val_agreement_rep,
                                                                       written_off=val_written_off,
                                                                       date_terminated=val_date_terminated,
                                                                       reason=val_reason2
                                                                       )

        # elif ws.cell(row=index, column=4).value:
        #
        #     if ws.cell(row=index, column=4).value.strip() != '':
        #         val_reason = ws.cell(row=index, column=4).value
        #     else:
        #         val_reason = 'TERMINATED - Worksheet = Live'

    for index in range(3, max_row3 + 1):

        wip_agreement_id = ws3.cell(index, 1).value

        if wip_agreement_id and wip_agreement_id != '':

            val_agreement_id = wip_agreement_id.strip()
            val_agreement_name = ws3.cell(row=index, column=2).value
            val_agreement_rep = ws3.cell(row=index, column=3).value

            if ws3.cell(row=index, column=5).value == 'Yes':
                val_written_off = True
            else:
                val_written_off = False

            val_date_terminated = ws3.cell(row=index, column=6).value
            if type(val_date_terminated) is not datetime:
                val_date_terminated = '2000-01-01'

            # write new row
            new_global_termination = ncf_global_terminations.objects.create(agreement_id=val_agreement_id,
                                                                       agreement_name=val_agreement_name,
                                                                       agreement_rep=val_agreement_rep,
                                                                       written_off=val_written_off,
                                                                       date_terminated=val_date_terminated,
                                                                       reason=val_reason3
                                                                       )

        # elif ws.cell(row=index, column=4).value:
        #
        #     if ws.cell(row=index, column=4).value.strip() != '':
        #         val_reason = ws.cell(row=index, column=4).value
        #     else:
        #         val_reason = 'TERMINATED - Worksheet = Live'

