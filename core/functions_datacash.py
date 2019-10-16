from openpyxl import load_workbook
from datetime import datetime
import pytz, datetime as ap_datetime

from .models import ncf_datacash_setups, ncf_datacash_drawdowns, go_extensions
from core_direct_debits.models import DDHistory


def app_process_datacash_drawdowns():

    filepath = "C:\/_go_file_interfaces_uat_staging/_go_datacash_interface/datacash_drawdowns.xlsx"
    wb = load_workbook(filepath, data_only=True)
    ws = wb['Sheet1']

    # get max row count
    max_row = ws.max_row + 1

    for index in range(2, max_row):

        val_dd_reference = ws.cell(row=index, column=1).value

        dd_history = DDHistory.objects.filter(reference=val_dd_reference).first()

        val_agreement_id = dd_history.agreement_no
        val_dd_setup_no = ws.cell(row=index, column=2).value
        val_dd_amount = ws.cell(row=index, column=3).value
        val_dd_method = ws.cell(row=index, column=4).value
        val_dd_request_date = ws.cell(row=index, column=5).value
        val_dd_batch_status = ws.cell(row=index, column=6).value
        val_dd_due_date = ws.cell(row=index, column=7).value
        val_dd_response = ws.cell(row=index, column=8).value
        val_dd_stage = ws.cell(row=index, column=9).value
        val_dd_bacs_reason = ws.cell(row=index, column=10).value

        val_dd_request_date = val_dd_request_date.replace(u'\xa0', ' ')
        struct_datetime_01 = pytz.utc.localize(datetime.strptime(val_dd_request_date, "%d %b %Y %H:%M:%S"))

        val_dd_due_date = val_dd_due_date.replace(u'\xa0', ' ')
        struct_datetime_02 = pytz.utc.localize(datetime.strptime(val_dd_due_date, "%d %b %Y"))

        val_max_due_datepre = ap_datetime.datetime(2001,1,1,1,1,1)
        val_max_due_date = pytz.utc.localize(val_max_due_datepre)

        if not ncf_datacash_drawdowns.objects.filter(dd_reference=val_dd_reference, dd_request_date=struct_datetime_01,
                                                dd_stage=val_dd_stage, dd_method=val_dd_method,
                                                dd_due_date=struct_datetime_02,
                                                dd_setup_no=val_dd_setup_no
                                                ).exists():

            new_drawdown = ncf_datacash_drawdowns.objects.create(agreement_id=val_agreement_id,
                                                            dd_reference=val_dd_reference,
                                                            dd_setup_no=val_dd_setup_no,
                                                            dd_amount=val_dd_amount,
                                                            dd_method=val_dd_method,
                                                            dd_request_date=struct_datetime_01,
                                                            dd_batch_status=val_dd_batch_status,
                                                            dd_due_date=struct_datetime_02,
                                                            dd_response=val_dd_response,
                                                            dd_stage=val_dd_stage,
                                                            dd_bacs_reason=val_dd_bacs_reason)

            if struct_datetime_02 > val_max_due_date:
                val_max_due_date = struct_datetime_02

            go_extension = go_extensions.objects.get(ap_extension_code='bacsrun')
            go_extension.ap_extension_last_interface_run = val_max_due_date
            go_extension.save()


def app_process_datacash_setups():

    filepath = "C:/_go_file_interfaces_uat_staging/_go_datacash_interface/datacash_setups.xlsx"
    wb = load_workbook(filepath, data_only=True)
    ws = wb['Sheet1']

    # get max row count
    max_row = ws.max_row + 1

    for index in range(2, max_row):

        val_dd_reference = ws.cell(row=index, column=1).value
        val_dd_reference = val_dd_reference.replace(" ", "")
        print(val_dd_reference)

        dd_history = DDHistory.objects.filter(reference=val_dd_reference).first()

        try:
            val_agreement_id = dd_history.agreement_no
        except:
            val_agreement_id = val_dd_reference

        val_account_name = ws.cell(row=index, column=2).value
        val_stage = ws.cell(row=index, column=3).value
        val_method = ws.cell(row=index, column=4).value
        val_request_date = ws.cell(row=index, column=5).value
        val_batch_status = ws.cell(row=index, column=6).value

        val_request_date = val_request_date.replace(u'\xa0', ' ')
        struct_datetime_01 = pytz.utc.localize(datetime.strptime(val_request_date, "%d %b %Y %H:%M:%S"))


        if not ncf_datacash_setups.objects.filter(agreement_id=val_agreement_id, dd_reference=val_dd_reference, dd_request_date=struct_datetime_01,
                                                dd_stage=val_stage, dd_method=val_method
                                                ).exists():

            new_setup = ncf_datacash_setups.objects.create(agreement_id=val_agreement_id,
                                                            dd_reference=val_dd_reference,
                                                            dd_account_name=val_account_name,
                                                            dd_method=val_method,
                                                            dd_request_date=struct_datetime_01,
                                                            dd_batch_status=val_batch_status,
                                                            dd_stage=val_stage
                                                            )
